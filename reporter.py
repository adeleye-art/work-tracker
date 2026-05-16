import sys
import os
import sqlite3
import requests
from datetime import date

# Resolve context_scraper when run from launchd (cwd may not be ~/work-ai)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from context_scraper import get_git_activity

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = os.path.expanduser("~/work-ai/logs.db")
REPORTS_DIR = os.path.expanduser("~/Documents/WorkReports")
POLL_SECONDS = 10

# Colours
C_DARK   = "2E4057"
C_TEAL   = "048A81"
C_GREEN  = "D4EDDA"
C_YELLOW = "FFF3CD"
C_BORDER = "DDDDDD"
C_WHITE  = "FFFFFF"


# ── helpers ───────────────────────────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(bottom=s)


def style_header(cell, bg=C_DARK):
    cell.font = Font(bold=True, color=C_WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def samples_to_hours(n):
    hours = n * POLL_SECONDS / 3600
    # Use enough decimal places so small values don't display as 0
    if hours < 0.01:
        return round(hours, 4)
    if hours < 0.1:
        return round(hours, 3)
    return round(hours, 2)


def categorize(app):
    n = app.lower()
    if any(x in n for x in ["code", "rider", "idea", "webstorm", "pycharm", "intellij", "goland"]):
        return "IDE"
    if "teams" in n:
        return "Teams"
    if any(x in n for x in ["chrome", "firefox", "safari", "edge", "arc", "brave"]):
        return "Browser"
    if any(x in n for x in ["terminal", "iterm", "warp"]):
        return "Terminal"
    return "Other"


# ── data layer ────────────────────────────────────────────────────────────────

def ensure_schema(conn):
    """Add new columns to existing DBs that predate the updated logger."""
    for col_def in ["project TEXT", "is_meeting INTEGER DEFAULT 0"]:
        try:
            conn.execute(f"ALTER TABLE activity ADD COLUMN {col_def}")
            conn.commit()
        except sqlite3.OperationalError:
            pass


def fetch_activity(target_date):
    conn = sqlite3.connect(DB_PATH)
    ensure_schema(conn)
    c = conn.cursor()
    c.execute("""
        SELECT app_name,
               COALESCE(project, 'Unknown') AS project,
               is_meeting,
               COUNT(*)        AS samples,
               MIN(timestamp)  AS first_seen,
               MAX(timestamp)  AS last_seen
        FROM activity
        WHERE date(timestamp) = ?
        GROUP BY app_name, project, is_meeting
        ORDER BY samples DESC
    """, (target_date,))
    rows = c.fetchall()
    conn.close()
    return rows


# ── Ollama AI summary ─────────────────────────────────────────────────────────

def ai_narrative(app_summary_text, git_summary_text):
    prompt = f"""You are a professional assistant for a Senior Full Stack Engineer.
Using the data below, write a concise, professional narrative summary of their workday.
Include: projects worked on, hours spent, meeting time, key code changes.
Keep it to 3-5 sentences. Do not use bullet points — write flowing prose.

APP ACTIVITY:
{app_summary_text}

GIT ACTIVITY:
{git_summary_text}"""

    try:
        resp = requests.post(
            "http://localhost:11434/api/generate",
            json={"model": "llama3.2:3b", "prompt": prompt, "stream": False},
            timeout=120,
        )
        return resp.json().get("response", "").strip()
    except Exception as e:
        return f"[AI summary unavailable — is Ollama running? Error: {e}]"


# ── Excel builder ─────────────────────────────────────────────────────────────

def generate_report(target_date=None):
    if target_date is None:
        target_date = date.today().isoformat()

    rows = fetch_activity(target_date)
    if not rows:
        print(f"No activity found for {target_date}. Is logger.py running?")
        return None

    # Aggregate
    detail_rows = []
    project_hours = {}
    category_hours = {}
    total_hours = 0.0
    meeting_hours = 0.0

    for app, project, is_meeting, samples, first_seen, last_seen in rows:
        hours = samples_to_hours(samples)
        cat = "Meeting" if is_meeting else categorize(app)
        detail_rows.append((app, project, cat, hours, first_seen, last_seen))
        project_hours[project] = project_hours.get(project, 0) + hours
        category_hours[cat] = category_hours.get(cat, 0) + hours
        total_hours += hours
        if is_meeting:
            meeting_hours += hours

    app_summary_text = "\n".join(
        f"- {app} | {project} | {'Meeting' if m else 'Work'} | {samples_to_hours(s):.2f}h"
        for app, project, m, s, _, _ in rows
    )
    # Only fetch git activity for projects the tracker saw you use today
    tracked_projects = {project for _, project, _, _, _, _ in rows}
    git_summary_text = get_git_activity(active_projects=tracked_projects, target_date=target_date)

    os.makedirs(REPORTS_DIR, exist_ok=True)
    path = os.path.join(REPORTS_DIR, f"{target_date}.xlsx")
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    # Title row
    ws["A1"] = f"Daily Work Report  —  {target_date}"
    ws["A1"].font = Font(bold=True, size=15)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    ws.append([])

    # Totals block
    for label, val in [
        ("Total Hours Tracked",   round(total_hours, 2)),
        ("Meeting Hours",         round(meeting_hours, 2)),
        ("Coding / Focus Hours",  round(total_hours - meeting_hours, 2)),
    ]:
        ws.append([label, val])
        r = ws.max_row
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2).font = Font(bold=True)

    ws.append([])

    # Detail header
    headers = ["App", "Project", "Category", "Hours", "First Active", "Last Active"]
    ws.append(headers)
    hr = ws.max_row
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(hr, col))

    for app, project, cat, hours, first_seen, last_seen in detail_rows:
        ws.append([app, project, cat, hours, first_seen, last_seen])
        r = ws.max_row
        fill = C_YELLOW if cat == "Meeting" else (C_GREEN if cat == "IDE" else None)
        for col in range(1, 7):
            cell = ws.cell(r, col)
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="left")
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    # ── Sheet 2: By Project ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Project")
    ws2.append(["Project", "Hours", "% of Day"])
    for col in range(1, 4):
        style_header(ws2.cell(1, col), C_TEAL)
    ws2.row_dimensions[1].height = 22

    for project, hours in sorted(project_hours.items(), key=lambda x: -x[1]):
        pct = f"{hours / total_hours * 100:.1f}%" if total_hours else "0%"
        ws2.append([project, round(hours, 2), pct])
        r = ws2.max_row
        for col in range(1, 4):
            ws2.cell(r, col).border = thin_border()

    ws2.append([])
    ws2.append(["TOTAL", round(total_hours, 2), "100%"])
    for col in range(1, 4):
        ws2.cell(ws2.max_row, col).font = Font(bold=True)

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12

    # ── Sheet 3: Git Activity ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Git Activity")
    ws3["A1"] = "Git Activity Summary"
    ws3["A1"].font = Font(bold=True, size=13)
    ws3.column_dimensions["A"].width = 100
    ws3.append([])
    for line in git_summary_text.split("\n"):
        ws3.append([line])

    # ── Sheet 4: AI Summary ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("AI Summary")
    ws4["A1"] = "AI-Generated Day Summary"
    ws4["A1"].font = Font(bold=True, size=13)
    ws4.column_dimensions["A"].width = 120

    print("Asking Ollama AI to summarise your day...")
    narrative = ai_narrative(app_summary_text, git_summary_text)

    ws4.append([])
    for line in narrative.split("\n"):
        ws4.append([line])
        ws4.cell(ws4.max_row, 1).alignment = Alignment(wrap_text=True)

    wb.save(path)
    print(f"Report saved: {path}")
    return path


if __name__ == "__main__":
    generate_report()
