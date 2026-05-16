import os
import sqlite3
import requests
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import config
from .scraper import get_git_activity

C_DARK   = "2E4057"
C_TEAL   = "048A81"
C_GREEN  = "D4EDDA"
C_YELLOW = "FFF3CD"
C_BORDER = "DDDDDD"
C_WHITE  = "FFFFFF"


def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(bottom=s)


def style_header(cell, bg=C_DARK):
    cell.font = Font(bold=True, color=C_WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def samples_to_hours(n, poll_seconds):
    hours = n * poll_seconds / 3600
    if hours < 0.01:
        return round(hours, 4)
    if hours < 0.1:
        return round(hours, 3)
    return round(hours, 2)


def categorize(app):
    n = app.lower()
    if any(x in n for x in ["code", "visual studio", "rider", "idea", "webstorm", "pycharm", "intellij", "goland"]):
        return "IDE"
    if any(x in n for x in ["teams", "zoom", "webex", "skype", "slack", "discord"]):
        return "Teams"
    if any(x in n for x in ["chrome", "firefox", "safari", "edge", "arc", "brave", "opera", "vivaldi"]):
        return "Browser"
    if any(x in n for x in ["terminal", "iterm", "warp"]):
        return "Terminal"
    return "Other"


def ensure_schema(conn):
    for col_def in ["project TEXT", "is_meeting INTEGER DEFAULT 0"]:
        try:
            conn.execute(f"ALTER TABLE activity ADD COLUMN {col_def}")
            conn.commit()
        except sqlite3.OperationalError:
            pass


def fetch_activity(db_path, target_date):
    conn = sqlite3.connect(db_path)
    ensure_schema(conn)
    c = conn.cursor()
    c.execute("""
        SELECT app_name,
               COALESCE(project, 'Unknown') AS project,
               is_meeting,
               COUNT(*)        AS samples,
               MIN(timestamp)  AS first_seen,
               MAX(timestamp)  AS last_seen
        FROM activity
        WHERE date(timestamp) = ?
        GROUP BY app_name, project, is_meeting
        ORDER BY samples DESC
    """, (target_date,))
    rows = c.fetchall()
    conn.close()
    return rows


def ai_narrative(app_summary, git_summary, cfg):
    if not cfg["ai"]["ollama_enabled"]:
        return "[AI summary disabled — set ollama_enabled = true in config to enable]"

    prompt = f"""You are a professional assistant for a Senior Software Engineer.
Using the data below, write a concise professional narrative summary of their workday.
Include: projects worked on, hours spent, meeting time, key commits.
3-5 sentences, flowing prose, no bullet points.

APP ACTIVITY:
{app_summary}

GIT ACTIVITY:
{git_summary}"""

    try:
        resp = requests.post(
            cfg["ai"]["ollama_url"] + "/api/generate",
            json={"model": cfg["ai"]["ollama_model"], "prompt": prompt, "stream": False},
            timeout=120,
        )
        return resp.json().get("response", "").strip()
    except Exception as e:
        return f"[AI summary unavailable — is Ollama running? Error: {e}]"


def generate_report(target_date=None):
    cfg          = config.load()
    db_path      = cfg["paths"]["db_path"]
    reports_dir  = cfg["paths"]["reports_dir"]
    poll_seconds = cfg["tracker"]["poll_interval"]

    if target_date is None:
        target_date = date.today().isoformat()

    rows = fetch_activity(db_path, target_date)
    if not rows:
        print(f"No activity found for {target_date}. Is the tracker running?")
        return None

    detail_rows    = []
    project_hours  = {}
    total_hours    = 0.0
    meeting_hours  = 0.0

    for app, project, is_meeting, samples, first_seen, last_seen in rows:
        hours = samples_to_hours(samples, poll_seconds)
        cat   = "Meeting" if is_meeting else categorize(app)
        detail_rows.append((app, project, cat, hours, first_seen, last_seen))
        project_hours[project] = project_hours.get(project, 0) + hours
        total_hours   += hours
        if is_meeting:
            meeting_hours += hours

    app_summary = "\n".join(
        f"- {app} | {proj} | {'Meeting' if m else 'Work'} | {samples_to_hours(s, poll_seconds):.2f}h"
        for app, proj, m, s, _, _ in rows
    )
    tracked_projects = {proj for _, proj, _, _, _, _ in rows}
    git_summary = get_git_activity(active_projects=tracked_projects, target_date=target_date)

    os.makedirs(reports_dir, exist_ok=True)
    path = os.path.join(reports_dir, f"{target_date}.xlsx")
    wb   = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Daily Work Report  —  {target_date}"
    ws["A1"].font = Font(bold=True, size=15)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    ws.append([])
    for label, val in [
        ("Total Hours Tracked",  round(total_hours, 2)),
        ("Meeting Hours",        round(meeting_hours, 2)),
        ("Coding / Focus Hours", round(total_hours - meeting_hours, 2)),
    ]:
        ws.append([label, val])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
        ws.cell(ws.max_row, 2).font = Font(bold=True)

    ws.append([])
    headers = ["App", "Project", "Category", "Hours", "First Active", "Last Active"]
    ws.append(headers)
    hr = ws.max_row
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(hr, col))

    for app, project, cat, hours, first_seen, last_seen in detail_rows:
        ws.append([app, project, cat, hours, first_seen, last_seen])
        r    = ws.max_row
        fill = C_YELLOW if cat == "Meeting" else (C_GREEN if cat == "IDE" else None)
        for col in range(1, 7):
            cell = ws.cell(r, col)
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal="left")
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)

    for col, w in zip("ABCDEF", [22, 30, 12, 10, 22, 22]):
        ws.column_dimensions[col].width = w

    # ── Sheet 2: By Project ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Project")
    ws2.append(["Project", "Hours", "% of Day"])
    for col in range(1, 4):
        style_header(ws2.cell(1, col), C_TEAL)
    ws2.row_dimensions[1].height = 22

    for proj, hours in sorted(project_hours.items(), key=lambda x: -x[1]):
        pct = f"{hours / total_hours * 100:.1f}%" if total_hours else "0%"
        ws2.append([proj, round(hours, 2), pct])
        for col in range(1, 4):
            ws2.cell(ws2.max_row, col).border = thin_border()

    ws2.append([])
    ws2.append(["TOTAL", round(total_hours, 2), "100%"])
    for col in range(1, 4):
        ws2.cell(ws2.max_row, col).font = Font(bold=True)
    for col, w in zip("ABC", [35, 12, 12]):
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: Git Activity ─────────────────────────────────────────────────
    ws3 = wb.create_sheet("Git Activity")
    ws3["A1"] = "Git Activity Summary"
    ws3["A1"].font = Font(bold=True, size=13)
    ws3.column_dimensions["A"].width = 100
    ws3.append([])
    for line in git_summary.split("\n"):
        ws3.append([line])

    # ── Sheet 4: AI Summary ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("AI Summary")
    ws4["A1"] = "AI-Generated Day Summary"
    ws4["A1"].font = Font(bold=True, size=13)
    ws4.column_dimensions["A"].width = 120
    ws4.append([])

    print("Asking Ollama AI to summarise your day...")
    narrative = ai_narrative(app_summary, git_summary, cfg)
    for line in narrative.split("\n"):
        ws4.append([line])
        ws4.cell(ws4.max_row, 1).alignment = Alignment(wrap_text=True)

    wb.save(path)
    print(f"Report saved: {path}")
    return path


if __name__ == "__main__":
    generate_report()
